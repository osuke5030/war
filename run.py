import requests
from bs4 import BeautifulSoup
import os
import json
import sys
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages")
from requests_oauthlib import OAuth1Session
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/lib/python3.10/site-packages")
from dotenv import find_dotenv, load_dotenv
import requests
import schedule
import time
import openpyxl


# .envファイルを探して読み込み
def job():
    try:
        env_file = find_dotenv()
        load_dotenv(env_file)  

        CONSUMER_KEY = os.environ.get('CONSUMER_KEY')
        CONSUMER_SECRET = os.environ.get('CONSUMER_SECRET')
        ACCESS_KEY = os.environ.get('ACCESS_KEY')
        ACCESS_KEY_SECRET = os.environ.get('ACCESS_KEY_SECRET')

        # Twitterの認証
        twitter = OAuth1Session(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_KEY, ACCESS_KEY_SECRET)
        print(twitter)

        #エンドポイント
        url_text = 'https://api.twitter.com/1.1/statuses/update.json'
        url_media = "https://upload.twitter.com/1.1/media/upload.json"
        # ここまでTwitter投稿の準備
        print("投稿準備完了")
        #スニだんのページの指定
        URL = 'https://sneakerwars.jp'
        # リクエストヘッダの指定
        headers = {"User-Agent": "hoge"}
        response = requests.get(URL,  headers=headers)
        r_text=response.text
        soup = BeautifulSoup(r_text, 'html.parser')
        print('スニのウォーの取得完了')

        soup_div=soup.find_all("div",attrs={"class","col-6"})[0]
        url=soup_div.find("a")['href']
        soup_url="https://sneakerwars.jp"+url
        soup_img=soup_div.find("a").find("img")['src']
        # ここから詳細ページ
        URL = soup_url
        # リクエストヘッダの指定
        headers = {"User-Agent": "hoge"}
        response = requests.get(URL,  headers=headers)
        r_text=response.text
        soup = BeautifulSoup(r_text, 'html.parser')
        print('詳細ページの取得完了')
        soup_name=soup.find("h1",attrs={"class","title"}).text
        response = requests.get(soup_img)
        image = response.content
        files = {"media" : image}
        req_media = twitter.post(url_media, files = files)
        media_id = json.loads(req_media.text)['media_id']
        print('画像の取得完了')
        soup_price=soup.find("dl",attrs={"class","extramode"}).text[soup.find("dl",attrs={"class","extramode"}).text.find("国内価格"):]
        params = {'status':"新スニ情報！！！!!!\n\n{}\n\n{}".format(soup_name,soup_price,),'media_ids':[media_id]}
        wb = openpyxl.load_workbook('war.xlsx')
        time.sleep(1)
        ws = wb["Sheet1"]
        for i in range(wb['Sheet1'].max_row):
            if ws.cell(row=i+1,column=1).value==params["status"]:
                print("投稿済みです")
                break    
            elif i==wb['Sheet1'].max_row-1:
                twitter.post(url_text, params = params)
                print("投稿しました")    
                ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                wb.save('war.xlsx')
            print("")
        print("処理終了しました")

    except IndexError:
        print("INDEX エラーです")       
        print("")

    except FileNotFoundError:
        print("NOT FILE エラーです")       
        print("")


def main():
    schedule.every(3).seconds.do(job)
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    main()